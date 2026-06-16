from pathlib import Path
import openpyxl as ax


root_dir = Path(__file__).parent / 'cases.xlsx'

def save_cases(client, note, status, deadline):
    if not root_dir.exists():
        wb = ax.Workbook()
        ws = wb.active
        ws.append(['Client', 'Note', 'Status', 'Deadline'])
        wb.save(root_dir)
        wb.close()

    wb = ax.load_workbook(root_dir)
    ws = wb.active
    deadline_str = deadline.toString('yyyy-MM-dd')
    ws.append([client, note, status, deadline_str])
    wb.save(root_dir)
    wb.close()

def load_cases():
    if not root_dir.exists():
        return []

    wb = ax.load_workbook(root_dir)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    return rows

def update_case(row, column, new_value):
    if not root_dir.exists():
        raise FileNotFoundError(f'Excel file not found: {root_dir}')

    wb = ax.load_workbook(root_dir)
    ws = wb.active
    ws.cell(row=row + 2, column=column + 1).value = new_value
    wb.save(root_dir)
    wb.close()
